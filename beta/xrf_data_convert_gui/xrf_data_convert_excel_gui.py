# coding=utf-8

import openpyxl as op
import re
import tkinter as tk


# Convert numeric columns to alphabet columns.
def numtochar(index):
    if index < 1:
        raise ValueError("Index is too small")
    result = ""
    while True:
        if index > 26:
            index, r = divmod(index - 1, 26)
            result = chr(r + ord('A')) + result
        else:
            return chr(index + ord('A') - 1) + result


# Read files in batches.
def xrf_data_convert_excel(filename_temp):
    b1_handler()

    # Main part.
    for file in file_list:
        # Read data and mark the active page as 'origin'.
        wb = op.load_workbook(file)
        origin = wb.active
        print("\n" + str(file) + " loads successfully.")
        # Read the num of target rows and columns, and record them as row, col.
        row = origin.cell(row=1, column=1).value
        col = origin.cell(row=2, column=1).value
        row = re.findall("\d+", row)[0]
        col = re.findall("\d+", col)[0]
        print('The number of target row is' + str(row))
        print('The number of target column is' + str(col))
        # Detect element types and store them in element[].
        element = []
        num = 0
        column = 6
        item = ''
        flag = True
        while flag:
            element.append(origin.cell(row=5, column=column).value)
            if element[num] == 'TIME':
                flag = False
            else:
                num += 1
                column += 1
                item += str(element[num - 1]) + ' '
        temp = 0
        for b in range(8, 8 + num):
            origin.cell(row=22, column=b).value = element[temp]
            temp = temp + 1
        print("The detected items are: " + str(item))
        print("Total " + str(num) + " items")
        print("Data is converting...")
        # Sheet creating.
        for i in range(0, num):
            temp = 'ws' + str(i)
            temp = wb.create_sheet(element[i], i + 1)
            # Copy the original data to the new sheet.
            sheets = wb.get_sheet_names()
            sheet = sheets[i + 1]
            item = wb.get_sheet_by_name(sheet)
            a = 1
            b = 23
            flag = True
            while flag:
                item.cell(row=a, column=1).value = origin.cell(row=b, column=(8 + i)).value
                if origin.cell(row=(b + 1), column=(8 + i)).value is None:
                    flag = False
                else:
                    a = a + 1
                    b = b + 1
            # Calculating for the new sheet.
            i = int(row) + 1
            j = int(col) + 2
            for a in range(1, int(i)):
                for b in range(2, int(j)):
                    c = numtochar(b - 1)
                    d = a
                    e = '=INDIRECT("a"&ROW(' + str(c) + str(d) + ')+(COLUMN(' + str(c) + str(d) + ')-1)*' + str(
                        row) + ')'
                    item.cell(row=a, column=b).value = e
        # Save file.
        new_filename = 'new_' + str(file)
        wb.save(filename=new_filename)
        print("Data converts successfully with a new filename " + new_filename)
    print("Finished.")


def b1_handler():
    b1_var.set('Processing')


root = tk.Tk()
root.title('xrf_data_convert_excel v3.0')
root.geometry('330x110')

l1 = tk.Label(root, text='Please enter the filename without ".xlsx" in turn.\nEnter "q()" to quit.')
l1.grid(row=1, column=0)


e1 = tk.Entry(root)
e1.grid(row=0, column=0)
file_list = []
filename_temp = e1.get()
filename = filename_temp + '.xlsx'
file_list.append(filename)


b1_var = tk.StringVar()
b1_var.set('Start')
b1 = tk.Button(root, textvariable=b1_var, command=xrf_data_convert_excel(file_list))
b1.grid(row=2, column=0)

root.mainloop()
