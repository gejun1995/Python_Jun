# -*- coding:utf-8 -*
# version 1.0

import openpyxl
import re

from openpyxl import load_workbook


# 数字列转换为字母列
def numtochar(idx):
    if idx < 1:
        raise ValueError("Index is too small")
    result = ""
    while True:
        if idx > 26:
            idx, r = divmod(idx - 1, 26)
            result = chr(r + ord('A')) + result
        else:
            return chr(idx + ord('A') - 1) + result


print "Version: Beta 1.0"
print '注意：需要先将".dat"格式文件转换成".xlsx"文件，并存储在本程序同一路径下！'

# 读取文件.
filename = str(raw_input('\n请输入文件名(无需".xlsx")!')) + '.xlsx'
wb = load_workbook(filename)

# 读取数据，并记活跃页为origin.
origin = wb.active
print "文件读取成功！"

# 读取目标转换行数、列数，并分别记为row、col.
row = origin.cell(row=1, column=1).value
col = origin.cell(row=2, column=1).value
row = re.findall("\d+", row)[0]
col = re.findall("\d+", col)[0]
print '\n目标行数为' + str(row)
print '目标列数为' + str(col)

# 检测元素种类，并存储至element[].
element = []
num = 0
column = 5
item = ''
flag = True
while flag:
    element.append(origin.cell(row=5, column=column).value)
    if element[num] == 'TIME':
        flag = False
    else:
        num += 1
        column += 1
        item += str(element[num - 1]) + ' '
temp = 0
for b in range(7, 7 + num):
    origin.cell(row=22, column=b).value = element[temp]
    temp = temp + 1
print "\n检测项目有：" + str(item)
print "共" + str(num) + "项"

# 创建sheet
print "\n正在进行数据转换……"
for i in range(0, num):
    temp = 'ws' + str(i)
    temp = wb.create_sheet(element[i], i + 1)
    # 将origin部分的目标项目数据复制到目标项目sheet中
    sheets = wb.get_sheet_names()
    sheet = sheets[i + 1]
    item = wb.get_sheet_by_name(sheet)
    a = 1
    b = 23
    flag = True
    while flag:
        item.cell(row=a, column=1).value = origin.cell(row=b, column=(7 + i)).value
        if origin.cell(row=(b + 1), column=(7 + i)).value == None:
            flag = False
        else:
            a = a + 1
            b = b + 1
    # 拉表
    i = int(row) + 1
    j = int(col) + 2
    for a in range(1, int(i)):
        for b in range(2, int(j)):
            c = numtochar(b - 1)
            d = a
            e = '=INDIRECT("a"&ROW(' + str(c) + str(d) + ')+(COLUMN(' + str(c) + str(d) + ')-1)*' + str(row) + ')'
            item.cell(row=a, column=b).value = e

 # 保存文件
filename = 'new_' + filename
wb.save(filename=filename)
print "\n数据转换成功！\n新文件名为：" + filename
