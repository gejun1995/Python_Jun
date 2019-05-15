import openpyxl as op

filename = 'G_A_cytoscape_LTCbrown and red.xlsx'
wb = op.load_workbook(filename)
sheet = wb.active

i = 2
count = 1
flag = True
while flag:
    a = str(sheet.cell(row=i, column=13).value)
    b = str(sheet.cell(row=i + 1, column=13).value)
    if a == b:
        count += 1
    else:
        for j in range(0, count):
            sheet.cell(row=i - j, column=22).value = count
        count = 1
    if a == 'None':
        flag = False
    else:
        i = i + 1

wb.save("M_" + filename)
