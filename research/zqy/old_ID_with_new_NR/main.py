import openpyxl as op
import linecache
import re

achn_excel = op.load_workbook(r'F:\TR_3\All_annotation.xlsx')
achn_sheet = achn_excel.active
print("open achn sheet")

acc_excel = op.load_workbook(r'F:\TR_3\Acc.xlsx')
acc_sheet = acc_excel.active
print("open acc sheet")

i = 2
flag = True
while flag:
    print("\n")
    print("-----------------------------------------------------------------")
    print("i = " + str(i))
    achn_excel_ID = str(achn_sheet.cell(row=i, column=1).value)
    print("achn_excel_ID is: " + achn_excel_ID)
    with open(r'F:\TR_3\Achn_gene.fa', 'r') as achn_db:
        for (achn_db_num, achn_db_ID) in enumerate(achn_db):
            achn_db_ID = re.findall(r'A\w*\d*', achn_db_ID)
            achn_db_ID = achn_db_ID[0]
            # print("compare with " + str(achn_db_ID[0]))
            if str(achn_db_ID) == str(achn_excel_ID):
                print("match!")
                seq = linecache.getline('F:\TR_3\Achn_gene.fa', achn_db_num + 2)
                seq = re.findall(r'[ATGC]*', seq)[0]
                print("seq is: " + seq)
                with open(r'F:\TR_3\seq.fa', 'a') as seqfa:
                    seqfa.write(">" + achn_db_ID)
                    seqfa.write("\n")
                    seqfa.write(seq)
                    seqfa.write("\n")
                with open(r'F:\TR_3\new_Acc_Gene.fa', 'r') as acc_db:
                    for (acc_db_num, acc_db_seq) in enumerate(acc_db):
                        acc_db_seq = re.findall(r'[ATGCatgc]*', acc_db_seq)[0]
                        acc_db_seq = acc_db_seq.upper()
                        # print("acc_db_seq is: " + str(acc_db_seq))
                        if str(acc_db_seq) == str(seq):
                            acc_db_ID = linecache.getline(r'F:\TR_3\new_Acc_Gene.fa', acc_db_num)
                            acc_db_ID = re.findall(r'A\w*\d*', acc_db_ID)[0]
                            print("acc_excel_ID is: " + acc_excel_ID)
                            for j in range(2, 35654):
                                acc_excel_ID = str(achn_sheet.cell(row=j, column=7).value)
                                if acc_db_ID == acc_excel_ID:
                                    acc_excel_NR = str(achn_sheet.cell(row=j, column=8).value)
                                    print("acc NR is: " + acc_excel_NR)
                                    achn_sheet.cell(row=i, column=13).value = acc_excel_NR
                                    achn_sheet.save("New_" + "All_annotation.xlsx")
                                    break
    if achn_excel_ID == 'None':
        flag = False
    else:
        i += 1

achn_sheet.save("New_" + "All_annotation.xlsx")
