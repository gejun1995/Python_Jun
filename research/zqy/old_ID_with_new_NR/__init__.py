import openpyxl as op
import linecache
import re

achn_excel = op.load_workbook(r'F:\TR_3\All_annotation.xls')
achn_sheet = achn_excel.active

acc_excel = op.load_workbook(r'F:\TR_3\Acc.xlsx')
acc_sheet = acc_excel.active

i = 2
flag = True
while flag:
    achn_excel_ID = str(achn_sheet.cell(row=i, column=1).value)
    achn_new_excel_ID = '>' + achn_excel_ID
    with open(r'F:\TR_3\Achn_gene.fa', 'r') as achn_db:
        for achn_db_ID in achn_db:
            if achn_db_ID == achn_new_excel_ID:
                seq = achn_db.next()
                with open(r'F:\TR_3\Acc_Gene.fa', 'r') as acc_db:
                    for (acc_db_num, acc_db_seq) in enumerate(acc_db):
                        if acc_db_seq == seq:
                            acc_db_ID = linecache.getline('F:\TR_3\Acc_Gene.fa',acc_db_num)
                            acc_excel_ID = acc_db_ID[1:]
                        for j in range(2, 35654):
                            acc_excel_ID = str(achn_sheet.cell(row=j, column=7).value)
                            if acc_db_ID == acc_excel_ID:
                                acc_excel_NR = str(achn_sheet.cell(row=j, column=8).value)
                                achn_sheet.cell(row=i, column=13).value = acc_excel_NR
                                break
if achn_excel_ID == 'None':
    flag = False
else:
    i += 1
