import openpyxl as op
import re

ori=[]
wb1=op.load_workbook('Zn-GSH.xlsx')
sheet1 = wb1.active
i=2
flag=True
while flag:
    a=str(sheet1.cell(row=i, column=1).value)
    ori.append(a)
    i+=1
    if a == 'None':
        flag=False

ori.pop()
for item in ori:
    print(item)

wb2 = op.load_workbook('new Zn-GSH.xlsx')
sheet2 = wb2.active
i =2
for item in ori:
    print(re.split(r'\s\s', str(item)))
    sep1 = re.split(r'\s\s', str(item))[0]
    sep2 = re.split(r'\s\s', str(item))[1]
    sheet2.cell(row=i, column=1).value = sep1
    sheet2.cell(row=i, column=2).value = sep2
    i=i+1

wb2.save('new Zn-GSh.xlsx')
