import openpyxl as op

name={}  #to store name and its value

wb1=op.load_workbook('species name.xlsx')
sheet1 = wb1.active
i=3  # the first row which species name starts
flag=True
while flag:
    a=str(sheet1.cell(row=i, column=1).value)
    b=str(sheet1.cell(row=i, column=2).value)
    name[a]=b
    i+=1
    if a == 'None':
        flag=False

wb2 = op.load_workbook('Network module separation and modularity calculation.xlsx')
sheet2 = wb2.active
i=4 # the first row which species name starts
flag = True
while flag:
    sheet2.cell(row=i, column=6).value=str(name[str(sheet2.cell(row=i,column=2).value)])
    i=i+1
    if str(sheet2.cell(row=i, column=2).value)=='None':
        flag=False
wb2.save('new '+'Network module separation and modularity calculation.xlsx')

print('Done')
