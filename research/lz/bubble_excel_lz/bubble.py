import openpyxl as op

wb = op.load_workbook('bubble.xlsx')
sheet_name_list = ["CKNS", "CKS", "HMNS", "HMS"]
position_name_list = ["Unplanted", "Bulk", "FR", "RS", "RP", "E"]
position_replica_dic = {"Unplanted": 3, "Bulk": 6, "FR": 6, "RS": 6, "RP": 6, "E": 6}
positioin_start_dic = {"Unplanted": 2, "Bulk": 5, "FR": 11, "RS": 17, "RP": 23, "E": 29}
otu_list = []  # todo

for sheet_name in sheet_name_list:
    sheet = wb.get_sheet_by_name(sheet_name)

    for position_name in position_name_list:
        filename = sheet_name + "_" + position_name + ".txt"
        with open(filename, 'w') as f:
            f.write(filename + "\n")
        title = "OTU_ID"
        for position_name_2 in position_name_list:
            for replica in range(1, position_replica_dic[position_name] + 1):
                title = title + "\t" + position_name_2 + str(replica)
        with open(filename, 'a') as f:
            f.write(title + "\n")

        i = 2
        traverse_flag = True
        while traverse_flag:
            otu_name = sheet.cell(row=i, column=1).value
            print("Current sheet name is:" + sheet_name)
            print("Position name is: " + position_name)
            print("Otu name is:" + otu_name + ". This is the " + str(i - 1) + " otu")

            otu_flag = True
            content = otu_name
            start = positioin_start_dic[position_name]
            end = positioin_start_dic[position_name] + position_replica_dic[position_name]
            for j in range(start, end):
                otu_value = sheet.cell(row=i, column=j).value
                content = content + "\t" + str(otu_value)
                if otu_value < 800:  # todo
                    otu_flag = False
            for j in range(2, start):
                otu_value = sheet.cell(row=i, column=j).value
                content = content + "\t" + str(otu_value)
                if otu_value > 100:  # todo
                    otu_flag = False
            for j in range(end, 35):
                otu_value = sheet.cell(row=i, column=j).value
                content = content + "\t" + str(otu_value)
                if otu_value > 100:  # todo
                    otu_flag = False
            if otu_flag is True:
                otu_list.append(otu_name)  # todo
                with open(filename, 'a') as f:
                    f.write(content + "\n")
                print("saved")
            i += 1
            if str(sheet.cell(row=i, column=1).value) == 'None':
                traverse_flag = False

print("Final:")  # todo
for item in otu_list:
    print(item, end='')
