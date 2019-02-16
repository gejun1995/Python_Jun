import re
import openpyxl as op

wb = op.load_workbook('HMMRhi2 0.920 edge_attribute.xlsx')
sheet = wb.active

name = []
i = 1
flag = True
while flag:
    name.append(str(sheet.cell(row=i, column=1).value))
    i += 1
    if str(sheet.cell(row=i, column=1).value) == 'None':
        flag = False
j = 1
for item in name:
    sep1 = re.findall(r'denovo\d+', item)[0]
    sep2 = re.findall(r'denovo\d+', item)[1]
    if item[-6] == '-':
        sep3 = item[-6:-4]
    else:
        sep3 = item[-5]

    sheet.cell(row=j, column=3).value = sep1
    sheet.cell(row=j, column=4).value = sep2
    sheet.cell(row=j, column=5).value = sep3
    j = j + 1

wb.save('new HMMRhi2 0.920 edge_attribute.xlsx')
