import openpyxl as op

# gap value
gap = -8

# read score for match or mismatch
filename = 'BLOSUM50.xlsx'
wb = op.load_workbook(filename)
sheet = wb.active

c = 2
flag = True
while (flag):
    score_p1 = sheet.cell(row=1, column=c + 1).value
    if score_p1 == None:
        flag = False
    else:
        c = c + 1
score_lenth = c
print("score lenth is: " + str(score_lenth))

score = {}
for c in range(2, score_lenth + 1):
    for r in range(2, score_lenth + 1):
        score_p1 = sheet.cell(row=1, column=c).value
        score_p2 = sheet.cell(row=r, column=1).value
        score_name = score_p1 + score_p2
        score_value = sheet.cell(row=r, column=c).value
        if score_value != None:
            score[score_name] = score_value
        score_name = score_p2 + score_p1
        if score_value != None:
            score[score_name] = score_value
print(score)

# for target two protein
wb = op.Workbook()
sheet = wb.active

P1 = "AGWGAHE"
P2 = "PAWEAEEG"

# write acid amino name
c = 4
r = 1
for i in range(len(P1)):
    sheet.cell(row=r, column=c).value = P1[i]
    c = c + 3

c = 1
r = 4
for i in range(len(P2)):
    sheet.cell(row=r, column=c).value = P2[i]
    r = r + 3

# score for match or mismatch
c = 4
r = 4
for i in range(len(P1)):
    for j in range(len(P2)):
        name1 = sheet.cell(row=1, column=c).value
        name2 = sheet.cell(row=r, column=1).value
        if name1 != None and name2 != None:
            name = name1 + name2
            sheet.cell(row=r, column=c).value = score[name]
        r = r + 3
    c = c + 3
    r = 4

# for extra information
sheet.cell(row=2, column=2).value = 0

c = 4
r = 4
for i in range(len(P1)):
    for j in range(len(P2)):
        name1 = sheet.cell(row=1, column=c).value
        name2 = sheet.cell(row=r, column=1).value
        if name1 != None and name2 != None:
            sheet.cell(row=2, column=c).value = gap
            sheet.cell(row=2, column=c - 1).value = "←"
            sheet.cell(row=2, column=c + 1).value = sheet.cell(row=2, column=c).value + sheet.cell(row=2,
                                                                                                   column=c - 2).value
            r = r + 3
    c = c + 3
    r = 4

c = 4
r = 4
for i in range(len(P1)):
    for j in range(len(P2)):
        name1 = sheet.cell(row=1, column=c).value
        name2 = sheet.cell(row=r, column=1).value
        if name1 != None and name2 != None:
            sheet.cell(row=r, column=2).value = gap
            sheet.cell(row=r - 1, column=2).value = "↑"
            sheet.cell(row=r + 1, column=2).value = sheet.cell(row=r, column=2).value + sheet.cell(row=r - 2,
                                                                                                   column=2).value
            r = r + 3
    c = c + 3
    r = 4

# score for gap
c = 4
r = 4
for i in range(len(P1)):
    for j in range(len(P2)):
        name1 = sheet.cell(row=1, column=c).value
        name2 = sheet.cell(row=r, column=1).value
        if name1 != None and name2 != None:
            sheet.cell(row=r, column=c + 1).value = gap
        r = r + 3
    c = c + 3
    r = 4

c = 4
r = 4
for i in range(len(P1)):
    for j in range(len(P2)):
        name1 = sheet.cell(row=1, column=c).value
        name2 = sheet.cell(row=r, column=1).value
        if name1 != None and name2 != None:
            sheet.cell(row=r + 1, column=c).value = gap
        r = r + 3
    c = c + 3
    r = 4

# for best hit
c = 4
r = 4
for i in range(len(P1)):
    for j in range(len(P2)):
        name1 = sheet.cell(row=1, column=c).value
        name2 = sheet.cell(row=r, column=1).value
        if name1 != None and name2 != None:
            top_left = sheet.cell(row=r, column=c).value
            top_right = sheet.cell(row=r, column=c + 1).value
            bottom_left = sheet.cell(row=r + 1, column=c).value
            top_left_value = sheet.cell(row=r - 2, column=c - 2).value + top_left
            top_right_value = sheet.cell(row=r - 2, column=c + 1).value + top_right
            bottom_left_value = sheet.cell(row=r + 1, column=c - 2).value + bottom_left
            if top_left_value >= top_right_value and top_left_value >= bottom_left_value:
                sheet.cell(row=r + 1, column=c + 1).value = top_left_value
                sheet.cell(row=r - 1, column=c - 1).value = "↖"
            if top_right_value >= top_left_value and top_right_value >= bottom_left_value:
                sheet.cell(row=r + 1, column=c + 1).value = top_right_value
                sheet.cell(row=r - 1, column=c + 1).value = "↑"
            if bottom_left_value >= top_left_value and bottom_left_value > top_right_value:
                sheet.cell(row=r + 1, column=c + 1).value = bottom_left_value
                sheet.cell(row=r + 1, column=c - 1).value = "←"
        r = r + 3
    c = c + 3
    r = 4

filename = 'align.xlsx'
wb.save(filename)
