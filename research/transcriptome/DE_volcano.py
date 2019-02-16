import openpyxl as op

filename = "uniport_nt_gene_annotated_RSEM.gene.counts.matrix.conditionA_vs_conditionB.DESeq2.DE_results.xlsx"
wb = op.load_workbook(filename)
sheet = wb.active

sheet.cell(row=1, column=15).value = "sig"

i = 2
while True:
    a = str(sheet.cell(row=i+1, column=1).value)

    log2FoldChange = float(sheet.cell(row=i, column=7).value)
    padj = float(sheet.cell(row=i, column=11).value)
    if padj < 0.001:
        if log2FoldChange > 2:
            sheet.cell(row=i, column=15).value = "down"
        elif log2FoldChange < -2:
            sheet.cell(row=i, column=15).value = "up"
        else:
            sheet.cell(row=i, column=15).value = "nochange"
    else:
        sheet.cell(row=i, column=15).value = "nochange"

    if a == 'None':
        break
    else:
        i = i + 1

wb.save("sig_" + filename)
