import openpyxl as op

filename = 'heatmap_DE.xlsx'
wb = op.load_workbook(filename)
sheet = wb.active

gene = []

i = 3
while True:
    a = str(sheet.cell(row=i, column=1).value)
    if a == "None":
        break
    gene.append(a)
    print(i)
    print(a)
    i = i + 1

print(gene)

filename = 'RSEM.gene.TMM.EXPR.matrix.xlsx'
wb = op.load_workbook(filename)
sheet = wb.active

i = 2
while True:
    a = str(sheet.cell(row=i, column=1).value)
    if a == "None":
        break
    if a in gene:
        print(a)
        sheet.cell(row=i, column=8).value = "change"
    i = i + 1
    print(i)

wb.save('heatmap ' + filename)
