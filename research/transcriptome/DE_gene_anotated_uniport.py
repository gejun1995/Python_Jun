import openpyxl as op
import re

filename = 'blastx.outfmt6.xlsx'
wb = op.load_workbook(filename)
sheet = wb.active

gene = {}

i = 2
while True:
    a = str(sheet.cell(row=i, column=1).value)
    pattern = re.compile(r"Cluster-\w*.\w*g\d")
    a = re.search(pattern, a)
    # print(type(a))
    if a is None:
        break
    a = a.group(0)
    b = str(sheet.cell(row=i, column=2).value)
    gene[a] = b
    i = i + 1
    print(a)
    print(gene[a])

filename = 'nt_gene_annotated_RSEM.gene.counts.matrix.conditionA_vs_conditionB.DESeq2.DE_results.xlsx'
wb = op.load_workbook(filename)
sheet = wb.active

i = 2
flag = True
while flag:
    a = str(sheet.cell(row=i, column=1).value)
    for key in gene.keys():
        if key == a:
            sheet.cell(row=i, column=13).value = gene[a]
    if a == 'None':
        flag = False
    else:
        i = i + 1

wb.save("uniport_" + filename)
