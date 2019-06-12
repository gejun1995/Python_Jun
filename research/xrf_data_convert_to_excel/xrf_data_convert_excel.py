# coding=utf-8

import openpyxl as op
import pandas as pd
import re

print(
"""
-----------------------------------------------------------------------------------------------------------------------------------------------------------
Purpose: Convert SSRL XRF data into '.xlsx' format.
Version 3.0

Please visit https://github.com/gejun1995/Python_Jun/blob/master/research/xrf_data_convert_to_excel/xrf_data_convert_excel.py to check for new updates.
Author information: Jun Ge, Zhejiang University, China. See at: www.gejun.me or www.gejunsci.com
Contact me: gejun@zju.edu.cn or gejun1995@gmail.com
All rights reserved.

By Jun Ge
Updated 6/12/2019
-----------------------------------------------------------------------------------------------------------------------------------------------------------
"""
)

def read_multi_files():
    filenames = []
    while True:
        filename = input('\nPlease input file name without extension.\nInput "s" to stop inputting and start converting.\n')
        if filename == 's':
            break
        filenames.append(filename)
    print("Inputted filenames:")
    for filename in filenames:
        print(filename)
    print("\n")

    return filenames

def dat_to_csv(filename):
    print("converting dat to csv...")
    with open(filename + ".dat", 'r') as file_dat:
        lines = file_dat.readlines()
    with open(filename + ".csv", 'w') as file_csv:
        for line in lines:
            line = re.sub(r'[\x09]', '\t', line)
            if line:
                file_csv.write(line)
    print("converted")

def csv_to_xlsx(filename):
    print("converting csv to xlsx...")
    csv = pd.read_csv(filename + '.csv', encoding='utf-8', sep='\t',dtype=object,low_memory=False)
    csv.to_excel(filename + '.xlsx', sheet_name='raw')
    wb = op.load_workbook(filename+".xlsx")
    raw = wb.active
    raw.delete_cols(1)
    print("converted")

def num_to_char(index):
    if index < 1:
        raise ValueError("Index is too small")
    result = ""
    while True:
        if index > 26:
            index, r = divmod(index - 1, 26)
            result = chr(r + ord('A')) + result
        else:
            return chr(index + ord('A') - 1) + result

def process(filename):
    # read row and column for each element
    print("-----------------------------------------------------------------------------------------------------------------------------------------------------------")
    print("opening file: " + filename +".xlsx")
    print("it may takes several minutes...")
    wb = op.load_workbook(filename+".xlsx")
    raw = wb.active
    print("\n" + str(filename) + " loads successfully.")
    row = raw.cell(row=1, column=1).value
    col = raw.cell(row=2, column=1).value
    row = re.findall("\d+", row)[0]
    col = re.findall("\d+", col)[0]
    print('The number of target row is ' + str(row))
    print('The number of target column is ' + str(col))

    # read element
    element = []
    element_num = 0
    element_column_1 = 2   # the column of element starts at row 5
    total_element_name = ''
    flag = True
    while flag:
        element.append(raw.cell(row=5, column=element_column_1).value)
        if element[element_num] == 'TIME':
            flag = False
        else:
            total_element_name += str(element[element_num]) + ' '
            element_num += 1
            element_column_1 += 1

    # paste element name on row 23
    element_num_temp = 0
    element_column_2 = 4 # the column of element starts at row 23
    for column_position in range(element_column_2, element_column_2 + element_num):
        raw.cell(row=22, column=column_position).value = element[element_num_temp]
        element_num_temp = element_num_temp + 1
    print("The detected elements are: " + str(total_element_name))
    print("Total " + str(element_num) + " elements")
    print("Data is converting...")

    # for each element
    for i in range(0, element_num):
        # create sheet for current element
        print("Current processing element is: " + element[i])
        print("Create new sheet for element...")
        wb.create_sheet(element[i], i + 1)

        # Copy the raw data of current element to the current sheet.
        print("Copy raw data to element sheet...")
        current_sheet = wb.get_sheet_by_name(element[i])
        a = 1
        b = 23
        flag = True
        while flag:
            current_sheet.cell(row=a, column=1).value = raw.cell(row=b, column=(element_column_2 + i)).value
            if raw.cell(row=(b + 1), column=(element_column_2 + i)).value is None:
                flag = False
            else:
                a = a + 1
                b = b + 1

        # Calculating for the current sheet.
        print("Calculate for element...")
        i = int(row) + 1
        j = int(col) + 2
        for a in range(1, int(i)):
            for b in range(2, int(j)):
                c = num_to_char(b - 1)
                d = a
                e = '=INDIRECT("a"&ROW(' + str(c) + str(d) + ')+(COLUMN(' + str(c) + str(d) + ')-1)*' + str(row) + ')'
                current_sheet.cell(row=a, column=b).value = e
        print("-------------------------------")

    print("saving data...")
    new_filename = 'new_' + str(filename) +".xlsx"
    wb.save(filename=new_filename)
    print("Data converts successfully with a new filename " + new_filename)
    print("\n\n")

def main():
    filenames=read_multi_files()
    print("Auto tranfering .dat to .xlsx comsumes time and computer resourses, tranfer manually if possible.")
    check = input("Transfer automatically any way by inputting 'yes', others skip.\n")
    if check == "yes":
        print("start auto transer...")
        for filename in filenames:
            print("current filename is: " + filename)
            dat_to_csv(filename)
            csv_to_xlsx(filename)
    else:
        print("ignored\n")
    for filename in filenames:
        process(filename)
    print("Finished.")


if __name__ == '__main__':
    main()
